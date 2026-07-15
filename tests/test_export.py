"""
Tests para GET /api/export/excel — verifica que el reporte incluya
las columnas Nombre y Calle.
"""
import io, base64, pytest
from unittest.mock import patch, MagicMock
import app as app_module
import openpyxl

AUTH = {'Authorization': 'Basic ' + base64.b64encode(b'usuario_test:clave_test').decode()}


def fake_db(fetchall_value):
    conn = MagicMock()
    conn.__enter__ = MagicMock(return_value=conn)
    conn.__exit__ = MagicMock(return_value=False)
    cur = MagicMock()
    cur.__enter__ = MagicMock(return_value=cur)
    cur.__exit__ = MagicMock(return_value=False)
    cur.fetchall.return_value = fetchall_value
    conn.cursor.return_value = cur
    return conn


@pytest.fixture
def client():
    app_module.app.config['TESTING'] = True
    with app_module.app.test_client() as c:
        yield c


class TestExportExcelCalle:
    def test_export_incluye_columnas_nombre_y_calle(self, client):
        fila = {
            'id': 1, 'folio': 'ENT-0001', 'tipo': 'ENTRADA', 'fecha': '2026-07-15',
            'hora': '10:00', 'pga': 'ESTERITO', 'detalle': 'ESCOMBRO', 'origen': 'NEGOCIO',
            'nombre': 'Juan Perez', 'calle': 'Av. Reforma', 'colonia': 'Centro',
            'vehiculo': 'CARROMATO', 'placa': 'ABC-123', 'm3': 2.5, 'obs': '',
            'creado_en': '2026-07-15 10:00:00'
        }
        with patch('app.get_db', return_value=fake_db([fila])):
            r = client.get('/api/export/excel', headers=AUTH)
        assert r.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert 'Nombre' in headers
        assert 'Calle' in headers

        row2 = [c.value for c in ws[2]]
        assert 'Juan Perez' in row2
        assert 'Av. Reforma' in row2
