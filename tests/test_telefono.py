"""
Tests del campo "Teléfono de contacto": se guarda al crear un registro,
se devuelve en el autocompletado por placa y aparece en la exportación.
Es opcional (no bloquea el guardado si viene vacío).
"""
import base64, io
import pytest
from unittest.mock import patch, MagicMock
import app as app_module

AUTH_ADMIN = {'Authorization': 'Basic ' + base64.b64encode(b'admin_test:clave_admin').decode()}

ENTRADA_OK = {'tipo': 'ENTRADA', 'fecha': '2026-08-03', 'pga': 'ESTERITO',
              'origen': 'NEGOCIO', 'placa': 'ABC-123-D', 'm3': 2.5}


def fake_db(fetchone_value=None, fetchall_value=None):
    conn = MagicMock()
    conn.__enter__ = MagicMock(return_value=conn)
    conn.__exit__ = MagicMock(return_value=False)
    cur = MagicMock()
    cur.__enter__ = MagicMock(return_value=cur)
    cur.__exit__ = MagicMock(return_value=False)
    cur.fetchone.return_value = fetchone_value
    cur.fetchall.return_value = fetchall_value or []
    conn.cursor.return_value = cur
    return conn


@pytest.fixture
def client():
    app_module.app.config['TESTING'] = True
    with app_module.app.test_client() as c:
        yield c


def test_crear_registro_guarda_telefono(client):
    """El INSERT debe incluir la columna telefono con el valor recibido."""
    conn = fake_db(fetchone_value={'valor': '1'})
    cur = conn.cursor.return_value
    with patch('app.get_db', return_value=conn):
        r = client.post('/api/registros',
                        json={**ENTRADA_OK, 'telefono': '8711234567'},
                        headers=AUTH_ADMIN)
    assert r.status_code == 201
    inserts = [c for c in cur.execute.call_args_list if 'INSERT INTO registros' in c.args[0]]
    assert inserts, "No se ejecutó el INSERT"
    sql, params = inserts[0].args[0], inserts[0].args[1]
    assert 'telefono' in sql, "El INSERT debe incluir la columna telefono"
    assert '8711234567' in params, "El teléfono debe pasarse como parámetro"


def test_crear_registro_sin_telefono_sigue_ok(client):
    """El teléfono es opcional: sin él, el registro se guarda igual."""
    conn = fake_db(fetchone_value={'valor': '1'})
    with patch('app.get_db', return_value=conn):
        r = client.post('/api/registros', json=ENTRADA_OK, headers=AUTH_ADMIN)
    assert r.status_code == 201


def test_buscar_placa_selecciona_telefono(client):
    """El autocompletado por placa debe traer también el teléfono."""
    conn = fake_db()
    cur = conn.cursor.return_value
    cur.fetchone.return_value = {'vehiculo': 'Volteo', 'detalle': '', 'origen': 'NEGOCIO',
                                 'nombre': 'Juan', 'calle': '', 'colonia': '', 'telefono': '8711234567'}
    with patch('app.get_db', return_value=conn):
        r = client.get('/api/registros/buscar-placa?q=ABC123', headers=AUTH_ADMIN)
    assert r.status_code == 200
    selects = [c.args[0] for c in cur.execute.call_args_list if 'FROM registros' in c.args[0]]
    assert selects, "No se ejecutó el SELECT de autocompletado"
    assert 'telefono' in selects[0], "El SELECT de buscar-placa debe incluir telefono"
    assert r.get_json()['telefono'] == '8711234567'


def test_export_incluye_columna_telefono(client):
    """La exportación a Excel debe tener una columna de teléfono."""
    conn = fake_db(fetchall_value=[{
        'id': 1, 'folio': 'ENT-0001', 'tipo': 'ENTRADA', 'fecha': '2026-08-03',
        'hora': '09:00', 'pga': 'ESTERITO', 'detalle': '', 'origen': 'NEGOCIO',
        'nombre': 'Juan', 'calle': '', 'colonia': '', 'vehiculo': 'Volteo',
        'placa': 'ABC-123-D', 'm3': 2.5, 'obs': '', 'creado_en': '2026-08-03',
        'telefono': '8711234567'}])
    with patch('app.get_db', return_value=conn):
        r = client.get('/api/export/excel', headers=AUTH_ADMIN)
    assert r.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert any(h and 'Tel' in str(h) for h in headers), f"Falta columna de teléfono en {headers}"
    fila = [c.value for c in ws[2]]
    assert '8711234567' in [str(v) for v in fila], "El teléfono no aparece en la fila exportada"
