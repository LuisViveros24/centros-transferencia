from flask import Flask, request, jsonify, render_template, send_file, Response, g, make_response
import psycopg2
import psycopg2.extras
import os, io, re, hmac, base64, math, json
from datetime import datetime, date
from functools import wraps

app = Flask(__name__)

# ── Configuración ──────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL')
AUTH_USER    = os.environ.get('AUTH_USER')
AUTH_PASS    = os.environ.get('AUTH_PASS')
# Opcionales: si están configuradas, AUTH_USER queda restringido a captura
# y ADMIN_USER tiene acceso total. Si no, AUTH_USER conserva acceso total
# (migración segura: el deploy no rompe nada antes de configurarlas en Render).
ADMIN_USER   = os.environ.get('ADMIN_USER')
ADMIN_PASS   = os.environ.get('ADMIN_PASS')
# Rol de solo lectura (tablero para dirección/jefe). Si no se configuran,
# el rol viewer simplemente no existe (nadie puede entrar como viewer).
VIEWER_USER  = os.environ.get('VIEWER_USER')
VIEWER_PASS  = os.environ.get('VIEWER_PASS')

if not DATABASE_URL:
    raise RuntimeError('Falta la variable de entorno DATABASE_URL')
if not AUTH_USER or not AUTH_PASS:
    raise RuntimeError('Faltan las variables de entorno AUTH_USER o AUTH_PASS')

# ── Base de datos ──────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def init_db():
    """Solo para desarrollo local. En producción usar migrate_data.py."""
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute('''
                    CREATE TABLE IF NOT EXISTS registros (
                        id        SERIAL PRIMARY KEY,
                        folio     TEXT NOT NULL,
                        tipo      TEXT NOT NULL,
                        fecha     DATE NOT NULL,
                        hora      TEXT,
                        pga       TEXT NOT NULL,
                        detalle   TEXT,
                        origen    TEXT,
                        nombre    TEXT,
                        telefono  TEXT,
                        calle     TEXT,
                        colonia   TEXT,
                        vehiculo  TEXT,
                        placa     TEXT,
                        m3        REAL DEFAULT 0,
                        obs       TEXT,
                        creado_en TIMESTAMP DEFAULT NOW()
                    )
                ''')
                cur.execute('''
                    CREATE TABLE IF NOT EXISTS config (
                        clave TEXT PRIMARY KEY,
                        valor TEXT
                    )
                ''')
                cur.execute('''
                    CREATE TABLE IF NOT EXISTS colonias_geo (
                        colonia_norm TEXT PRIMARY KEY,
                        lat          REAL,
                        lng          REAL,
                        estado       TEXT NOT NULL,
                        creado_en    TIMESTAMP DEFAULT NOW()
                    )
                ''')
                cur.execute('''
                    CREATE TABLE IF NOT EXISTS domicilios (
                        id              SERIAL PRIMARY KEY,
                        folio           TEXT NOT NULL,
                        folio_acta      TEXT,
                        fecha           DATE NOT NULL,
                        direccion       TEXT,
                        uso             TEXT,
                        nombre_comercio TEXT,
                        estado          TEXT,
                        problematica    TEXT,
                        accion          TEXT,
                        equipo          TEXT,
                        plazo_horas     INTEGER,
                        lat             REAL,
                        lng             REAL,
                        obs             TEXT,
                        foto_pdf        BYTEA,
                        creado_en       TIMESTAMP DEFAULT NOW()
                    )
                ''')
                for _col, _tipo in (('folio_acta','TEXT'), ('accion','TEXT'),
                                    ('equipo','TEXT'), ('plazo_horas','INTEGER'),
                                    ('lat','REAL'), ('lng','REAL'),
                                    ('obs','TEXT'), ('foto_pdf','BYTEA'),
                                    ('cumplido','BOOLEAN'), ('cumplido_en','TIMESTAMP'),
                                    ('cumplido_obs','TEXT'), ('cumplido_por','TEXT'),
                                    ('multa','BOOLEAN'),
                                    ('canalizado_ingresos','BOOLEAN'),
                                    ('canalizado_en','TIMESTAMP'), ('canalizado_por','TEXT'),
                                    ('incumplimiento','BOOLEAN')):
                    cur.execute(f'ALTER TABLE domicilios ADD COLUMN IF NOT EXISTS {_col} {_tipo}')
                cur.execute(
                    "INSERT INTO config VALUES ('folio_base','1') ON CONFLICT DO NOTHING"
                )
                cur.execute(
                    "INSERT INTO config VALUES ('folio_dom','1') ON CONFLICT DO NOTHING"
                )
    finally:
        conn.close()

# ── Autenticación y roles ──────────────────────────────────────
def _cred_ok(auth, exp_user, exp_pass):
    if not exp_user or not exp_pass or not auth or auth.username is None or auth.password is None:
        return False
    # Comparar en bytes UTF-8: hmac.compare_digest sobre str truena con
    # caracteres no-ASCII (acentos, ñ), lo que rompería un login con esos
    # caracteres en la contraseña.
    def eq(a, b):
        return hmac.compare_digest(a.encode('utf-8'), b.encode('utf-8'))
    return eq(auth.username, exp_user) and eq(auth.password, exp_pass)

def _rol_de(auth):
    """'admin', 'viewer', 'captura' o None según las credenciales recibidas."""
    if _cred_ok(auth, ADMIN_USER, ADMIN_PASS):
        return 'admin'
    if _cred_ok(auth, VIEWER_USER, VIEWER_PASS):
        return 'viewer'
    if _cred_ok(auth, AUTH_USER, AUTH_PASS):
        # Sin admin configurado, el usuario de captura conserva acceso total
        return 'captura' if (ADMIN_USER and ADMIN_PASS) else 'admin'
    return None

def _respuesta_401():
    return Response(
        'Acceso restringido',
        401,
        {'WWW-Authenticate': 'Basic realm="CT App"'}
    )

def requiere_auth(f):
    """Captura o admin. El rol de solo lectura (viewer) queda excluido:
    no debe ver el formulario de captura, solo el tablero."""
    @wraps(f)
    def decorado(*args, **kwargs):
        rol = _rol_de(request.authorization)
        if rol is None:
            return _respuesta_401()
        if rol == 'viewer':
            return jsonify({'error': 'Cuenta de solo lectura. Usa el tablero en /tablero.'}), 403
        g.rol = rol
        return f(*args, **kwargs)
    return decorado

def requiere_admin(f):
    """Solo el usuario administrador."""
    @wraps(f)
    def decorado(*args, **kwargs):
        rol = _rol_de(request.authorization)
        if rol is None:
            return _respuesta_401()
        if rol != 'admin':
            return jsonify({'error': 'Se requiere usuario administrador'}), 403
        g.rol = rol
        return f(*args, **kwargs)
    return decorado

def requiere_tablero(f):
    """Tablero de solo lectura: accesible a admin o al rol viewer (jefe)."""
    @wraps(f)
    def decorado(*args, **kwargs):
        rol = _rol_de(request.authorization)
        if rol is None:
            return _respuesta_401()
        if rol not in ('admin', 'viewer'):
            return jsonify({'error': 'Acceso restringido al tablero'}), 403
        g.rol = rol
        return f(*args, **kwargs)
    return decorado

# ── Folio ──────────────────────────────────────────────────────
def next_folio(tipo, cur):
    """Genera el siguiente folio. Debe llamarse dentro de la misma transacción
    que el INSERT, para que el SELECT FOR UPDATE mantenga el lock."""
    cur.execute(
        "SELECT valor FROM config WHERE clave='folio_base' FOR UPDATE"
    )
    row = cur.fetchone()
    if row is None:
        raise RuntimeError("Falta la fila 'folio_base' en la tabla config")
    n = int(row['valor'])
    prefix = 'ENT-' if tipo == 'ENTRADA' else 'SAL-'
    folio = prefix + str(n).zfill(4)
    cur.execute(
        "UPDATE config SET valor=%s WHERE clave='folio_base'", (n + 1,)
    )
    return folio

def next_folio_dom(cur):
    """Siguiente folio de domicilios (prefijo DOM-). Debe llamarse dentro de la
    misma transacción que el INSERT, para mantener el lock del SELECT FOR UPDATE."""
    cur.execute(
        "SELECT valor FROM config WHERE clave='folio_dom' FOR UPDATE"
    )
    row = cur.fetchone()
    if row is None:
        raise RuntimeError("Falta la fila 'folio_dom' en la tabla config")
    n = int(row['valor'])
    folio = 'DOM-' + str(n).zfill(4)
    cur.execute(
        "UPDATE config SET valor=%s WHERE clave='folio_dom'", (n + 1,)
    )
    return folio

# ── Rutas ──────────────────────────────────────────────────────
@app.route('/')
@requiere_auth
def index():
    # no-store: el navegador siempre pide la versión más reciente del HTML.
    # Evita que un operador quede con una copia vieja en caché tras un deploy
    # (frontend viejo + backend nuevo = errores confusos).
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    return resp

@app.route('/logout')
def logout():
    """Cierre de sesión para HTTP Basic Auth. Responde 401 SIEMPRE (aunque
    lleguen credenciales válidas) con el mismo realm de la app, para que el
    navegador considere inválidas las credenciales cacheadas y vuelva a pedir
    usuario y contraseña en el siguiente acceso."""
    return Response(
        'Sesión cerrada. Para volver a entrar, abre de nuevo la aplicación.',
        401,
        {'WWW-Authenticate': 'Basic realm="CT App"', 'Cache-Control': 'no-store'}
    )

@app.route('/api/whoami', methods=['GET'])
@requiere_auth
def whoami():
    """Rol del usuario autenticado; el frontend adapta la navegación."""
    return jsonify({'rol': g.rol})

@app.route('/api/badges', methods=['GET'])
@requiere_auth
def badges():
    """Conteo de entradas/salidas de una fecha, para los badges del menú.
    Accesible a captura (a diferencia de GET /api/registros, que es admin)."""
    fecha = request.args.get('fecha', str(date.today()))
    try:
        datetime.strptime(fecha, '%Y-%m-%d')
    except (ValueError, TypeError):
        return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT COUNT(*) c FROM registros WHERE fecha=%s AND tipo='ENTRADA'", [fecha])
                ent = cur.fetchone()['c']
                cur.execute(
                    "SELECT COUNT(*) c FROM registros WHERE fecha=%s AND tipo='SALIDA'", [fecha])
                sal = cur.fetchone()['c']
                cur.execute(
                    "SELECT COUNT(*) c FROM domicilios WHERE fecha=%s", [fecha])
                dom = cur.fetchone()['c']
    finally:
        conn.close()
    return jsonify({'ent': ent, 'sal': sal, 'dom': dom})

@app.route('/api/registros', methods=['GET'])
@requiere_admin
def get_registros():
    fecha = request.args.get('fecha')
    tipo  = request.args.get('tipo')
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                q = "SELECT * FROM registros WHERE 1=1"
                params = []
                if fecha:
                    q += " AND fecha=%s"; params.append(fecha)
                if tipo:
                    q += " AND tipo=%s"; params.append(tipo)
                q += " ORDER BY id DESC"
                cur.execute(q, params)
                rows = cur.fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/registros', methods=['POST'])
@requiere_auth
def crear_registro():
    d = request.get_json()
    if not d:
        return jsonify({'error': 'JSON requerido'}), 400

    # m³ obligatorio y mayor a 0 (entradas y salidas)
    try:
        m3 = float(d.get('m3'))
    except (TypeError, ValueError):
        m3 = 0
    if m3 <= 0:
        return jsonify({'error': 'Los metros cúbicos son obligatorios y deben ser mayores a 0'}), 400

    # Placa obligatoria en entradas: 6-8 alfanuméricos (ignorando espacios
    # y guiones) o el valor especial SIN PLACAS. Regla validada contra los
    # datos de producción (94% de las placas tienen 7 alfanuméricos).
    if d.get('tipo', 'ENTRADA') == 'ENTRADA':
        placa_norm = re.sub(r'[^A-Z0-9]', '', str(d.get('placa') or '').upper())
        if placa_norm != 'SINPLACAS' and not (6 <= len(placa_norm) <= 8):
            return jsonify({'error': 'Placa inválida: escribe de 6 a 8 letras/números, o "SIN PLACAS"'}), 400

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                folio = next_folio(d.get('tipo', 'ENTRADA'), cur)
                cur.execute('''
                    INSERT INTO registros
                    (folio,tipo,fecha,hora,pga,detalle,origen,nombre,telefono,calle,colonia,vehiculo,placa,m3,obs)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ''', (
                    folio,
                    d.get('tipo', 'ENTRADA'),
                    d.get('fecha', str(date.today())),
                    d.get('hora', ''),
                    d.get('pga', ''),
                    d.get('detalle', ''),
                    d.get('origen', ''),
                    d.get('nombre', ''),
                    d.get('telefono', ''),
                    d.get('calle', ''),
                    d.get('colonia', ''),
                    d.get('vehiculo', ''),
                    d.get('placa', ''),
                    m3,
                    d.get('obs', '')
                ))
    finally:
        conn.close()
    return jsonify({'ok': True, 'folio': folio}), 201

@app.route('/api/registros/buscar-placa', methods=['GET'])
@requiere_auth
def buscar_placa():
    """Devuelve los datos del último registro de entrada con esa placa (parcial)."""
    q = request.args.get('q', '').strip().upper()
    if len(q) < 3:
        return jsonify(None)
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT vehiculo, detalle, origen, nombre, telefono, calle, colonia
                    FROM registros
                    WHERE UPPER(placa) LIKE %s AND tipo='ENTRADA'
                    ORDER BY id DESC LIMIT 1
                """, (q + '%',))
                row = cur.fetchone()
    finally:
        conn.close()
    return jsonify(dict(row) if row else None)

@app.route('/api/registros/<int:rid>', methods=['DELETE'])
@requiere_admin
def eliminar_registro(rid):
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM registros WHERE id=%s", (rid,))
                if cur.rowcount == 0:
                    return jsonify({'error': 'Registro no encontrado'}), 404
                # Recalcular folio_base al número más alto existente + 1
                # Si no quedan registros, vuelve a 1
                cur.execute("""
                    UPDATE config SET valor = (
                        SELECT COALESCE(MAX(SUBSTRING(folio FROM 5)::INTEGER), 0) + 1
                        FROM registros
                    ) WHERE clave = 'folio_base'
                """)
    finally:
        conn.close()
    return jsonify({'ok': True})

# ── Domicilios (control de predios) ────────────────────────────
# Catálogos de referencia. Se pueden ampliar/editar libremente; el backend
# no restringe los valores (para facilitar cambios futuros sin migración).
USOS_PREDIO    = ('Comercial', 'Habitacional', 'Ambos')
ESTADOS_PREDIO = ('Habitado', 'Deshabitado', 'Baldío')
PROBLEMATICAS  = ('Escombro', 'Basura', 'Enyerbado', 'Desecho Vegetal',
                  'Desecho Vegetal Ajeno', 'Poda de Árboles', 'Vehículo Chatarra',
                  'Descarga de grasa', 'Otro')
# Fundamento legal por problemática (verificado contra los reglamentos de Torreón).
# 'art' = artículo(s) que se infringe · 'tag': DS=Regl. Desarrollo Sustentable, Limp=Regl. Limpieza.
FUNDAMENTOS = {
    'Escombro':          {'art': 'Art. 270 fr. XXXII (base 209)', 'tag': 'DS'},
    'Basura':            {'art': 'Art. 270 fr. XXXIX/XL',         'tag': 'DS'},
    'Enyerbado':         {'art': 'Art. 13 fr. I / 16 fr. I',      'tag': 'Limp'},
    'Desecho Vegetal':   {'art': 'Art. 13 fr. VI / 270 fr. XLII', 'tag': 'Limp/DS'},
    'Desecho Vegetal Ajeno': {'art': 'Art. 13 fr. VI / 270 fr. XLII', 'tag': 'Limp/DS'},
    'Poda de Árboles':   {'art': 'Art. 209 / 270 fr. XLII',       'tag': 'DS'},
    'Vehículo Chatarra': {'art': 'Art. 200 y 209',                'tag': 'DS'},
    'Descarga de grasa': {'art': 'Art. 106 / 270 fr. XII',        'tag': 'DS'},
}
ACCIONES       = ('Notificado', 'Amonestado', 'Multado')
# Equipos de captura: nombres provisionales; se editarán cuando lleguen los reales.
EQUIPOS        = (
    'Equipo 1 · César Alvarado, Rafael Moisés y Cristina Estrada',
    'Equipo 2 · Alberto Adame Martínez e Itzel García',
    'Equipo 3 · José Guajardo, César Crispín y Dulce Pérez',
    'Equipo 4 · Abraham Álvarez y Luis Viveros',
    'Equipo 5 · Ernesto Escalera y Salvador García',
)
# Expresión SQL del datetime límite del plazo (misma lógica que _fecha_limite):
# plazo_horas 0 = 'el mismo día' (vence 23:59 del día de inspección); NULL = sin plazo.
SQL_LIMITE = ("CASE WHEN plazo_horas IS NULL THEN NULL "
              "WHEN plazo_horas = 0 THEN (fecha::timestamp + interval '23 hours 59 minutes') "
              "ELSE (fecha::timestamp + make_interval(hours => plazo_horas)) END")
# ── Operativo: polígonos por equipo (por defecto; editable en la tabla config) ──
POLY_ASIGNADOS = {
    '1': [2, 9, 15, 16, 20, 26, 31, 39], '2': [4, 6, 12, 17, 21, 27, 32, 40],
    '3': [1, 11, 14, 18, 28, 33, 34, 41], '4': [3, 8, 10, 19, 29, 30, 35, 42],
    '5': [5, 7, 13, 24, 25, 37, 36, 38],
}
POLY_CUBIERTOS = [1, 2, 3, 4, 5, 6]
POLY_COLORS = {'1': '#2980b9', '2': '#27ae60', '3': '#e67e22', '4': '#8e44ad', '5': '#e74c3c'}
MANZANAS_DEFAULT = {"1":8,"2":9,"3":7,"4":10,"5":7,"6":9,"7":6,"8":7,"9":9,"10":7,"11":7,"12":9,"13":7,"14":6,"15":6,"16":8,"17":6,"18":13,"19":14,"20":6,"21":8,"24":4,"25":7,"26":7,"27":8,"28":8,"29":8,"30":6,"31":8,"32":7,"33":7,"34":7,"35":7,"36":7,"37":8,"38":7,"39":6,"40":6,"41":6,"42":5}

def _poligonos_payload(pcfg):
    """Arma la config de polígonos para el tablero/PDF: asignados, manzanas totales,
    manzanas cubiertas por polígono, y la lista de polígonos completos (derivada)."""
    asign = json.loads(pcfg['poligonos_asignados']) if pcfg.get('poligonos_asignados') else POLY_ASIGNADOS
    mz = json.loads(pcfg['manzanas_por_poligono']) if pcfg.get('manzanas_por_poligono') else MANZANAS_DEFAULT
    if pcfg.get('manzanas_cubiertas'):
        mc = json.loads(pcfg['manzanas_cubiertas'])
    else:
        cub_list = json.loads(pcfg['poligonos_cubiertos']) if pcfg.get('poligonos_cubiertos') else POLY_CUBIERTOS
        mc = {str(p): mz.get(str(p), 0) for p in cub_list}
    completos = sorted(int(p) for p in mz if int(mc.get(p, 0) or 0) >= int(mz[p]) and int(mz[p]) > 0)
    return {'asignados': asign, 'manzanas': mz, 'cubiertas': mc, 'colores': POLY_COLORS, 'cubiertos': completos}
MAX_FOTOS      = 5                     # fotos por domicilio
MAX_FOTO_BYTES = 5 * 1024 * 1024       # 5 MB por foto (ya comprimida en el cliente)

def _fecha_limite(fecha, plazo_horas):
    """datetime límite = fecha de inspección + plazo. plazo_horas=0 significa
    'el mismo día' (vence a las 23:59 del día de inspección). None si no aplica."""
    if plazo_horas is None:
        return None
    try:
        ph = int(plazo_horas)
    except (ValueError, TypeError):
        return None
    from datetime import timedelta
    base = datetime.combine(fecha, datetime.min.time()) if isinstance(fecha, date) \
        else datetime.strptime(str(fecha), '%Y-%m-%d')
    if ph == 0:
        return base.replace(hour=23, minute=59)
    return base + timedelta(hours=ph)

def _plazo_texto(plazo_horas):
    """Texto legible del plazo para PDF/Excel."""
    if plazo_horas is None:
        return ''
    if plazo_horas == 0:
        return 'El mismo día'
    return f'{plazo_horas} horas'

def _fotos_a_pdf(fotos_dec, meta_lines):
    """Genera un PDF (bytes) con una portada de datos + una página por foto."""
    from PIL import Image, ImageDraw, ImageFont
    W, H = 1240, 1754  # ~A4 a 150 dpi
    cover = Image.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(cover)
    try:
        f_title = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 46)
        f_body  = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 32)
    except Exception:
        f_title = ImageFont.load_default()
        f_body  = ImageFont.load_default()
    draw.text((80, 80), 'Control de domicilios', font=f_title, fill='black')
    y = 180
    for line in meta_lines:
        draw.text((80, y), line, font=f_body, fill='black')
        y += 52
    pages = [cover]
    for _mime, raw in fotos_dec:
        try:
            im = Image.open(io.BytesIO(raw)).convert('RGB')
            pages.append(im)
        except Exception:
            continue
    buf = io.BytesIO()
    pages[0].save(buf, format='PDF', save_all=True,
                  append_images=pages[1:] if len(pages) > 1 else [])
    buf.seek(0)
    return buf.read()

def _decode_data_url(s):
    """Convierte 'data:image/jpeg;base64,AAAA' en (mime, bytes). Devuelve (None, None) si falla."""
    if not s or not isinstance(s, str) or ',' not in s:
        return None, None
    header, b64 = s.split(',', 1)
    mime = 'image/jpeg'
    m = re.match(r'data:([^;]+);base64', header)
    if m:
        mime = m.group(1)
    if mime not in ('image/jpeg', 'image/png', 'image/webp'):
        return None, None
    try:
        raw = base64.b64decode(b64)
    except Exception:
        return None, None
    if not raw or len(raw) > MAX_FOTO_BYTES:
        return None, None
    return mime, raw

@app.route('/api/domicilios', methods=['GET'])
@requiere_admin
def get_domicilios():
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # No devolvemos foto_pdf (bytea pesado); solo si existe.
                cur.execute("""
                    SELECT id, folio, folio_acta, fecha, direccion, uso, nombre_comercio,
                           estado, problematica, accion, equipo, plazo_horas, lat, lng, obs,
                           (foto_pdf IS NOT NULL) AS has_pdf, creado_en
                    FROM domicilios ORDER BY id DESC
                """)
                rows = cur.fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = dict(r)
        lim = _fecha_limite(r['fecha'], r.get('plazo_horas'))
        row['fecha_limite'] = lim.strftime('%Y-%m-%d %H:%M') if lim else ''
        out.append(row)
    return jsonify(out)

@app.route('/api/domicilios', methods=['POST'])
@requiere_auth
def crear_domicilio():
    d = request.get_json()
    if not d:
        return jsonify({'error': 'JSON requerido'}), 400

    fecha = d.get('fecha') or str(date.today())
    try:
        datetime.strptime(fecha, '%Y-%m-%d')
    except (ValueError, TypeError):
        return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400

    direccion = (d.get('direccion') or '').strip()
    if not direccion:
        return jsonify({'error': 'La dirección es obligatoria.'}), 400

    uso = (d.get('uso') or '').strip()
    if uso not in USOS_PREDIO:
        return jsonify({'error': 'Selecciona el uso del predio (Comercial o Habitacional).'}), 400

    nombre_comercio = (d.get('nombre_comercio') or '').strip()
    if uso == 'Comercial' and not nombre_comercio:
        return jsonify({'error': 'Para uso Comercial, el nombre del comercio es obligatorio.'}), 400
    if uso != 'Comercial':
        nombre_comercio = ''  # solo aplica a comercios

    estado       = (d.get('estado') or '').strip()
    equipo       = (d.get('equipo') or '').strip()
    accion       = (d.get('accion') or '').strip()
    folio_acta   = (d.get('folio_acta') or '').strip()
    obs          = (d.get('obs') or '').strip()

    # Problemática: puede venir como lista (selección múltiple) o texto.
    prob_in = d.get('problematica')
    if isinstance(prob_in, list):
        problematica = ', '.join(str(p).strip() for p in prob_in if str(p).strip())
    else:
        problematica = (prob_in or '').strip()

    # Coordenadas (geolocalización del celular), opcionales.
    def _num(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return None
    lat = _num(d.get('lat'))
    lng = _num(d.get('lng'))

    plazo_horas = d.get('plazo_horas')
    if plazo_horas in ('', None):
        plazo_horas = None
    else:
        try:
            plazo_horas = int(plazo_horas)   # 0 = el mismo día
            if plazo_horas < 0:
                plazo_horas = None
        except (ValueError, TypeError):
            return jsonify({'error': 'El plazo debe ser un número entero de horas (o 0 = el mismo día).'}), 400

    # Fotos: lista de data URLs (base64). Opcionales, máximo MAX_FOTOS.
    fotos_in = d.get('fotos') or []
    if not isinstance(fotos_in, list):
        fotos_in = []
    if len(fotos_in) > MAX_FOTOS:
        return jsonify({'error': f'Máximo {MAX_FOTOS} fotografías por domicilio.'}), 400
    fotos_dec = []
    for s in fotos_in:
        mime, raw = _decode_data_url(s)
        if raw is None:
            return jsonify({'error': 'Una de las fotografías no es válida o excede el tamaño permitido.'}), 400
        fotos_dec.append((mime, raw))

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                folio = next_folio_dom(cur)
                # Generar el PDF de fotos (con portada de datos) al momento de capturar.
                foto_pdf = None
                if fotos_dec:
                    coords = f'{lat}, {lng}' if (lat is not None and lng is not None) else '—'
                    meta = [
                        f'Folio captura: {folio}',
                        f'Folio acta física: {folio_acta or "—"}',
                        f'Fecha inspección: {fecha}',
                        f'Dirección: {direccion}',
                        f'Uso: {uso}' + (f' — {nombre_comercio}' if nombre_comercio else ''),
                        f'Estado: {estado or "—"}',
                        f'Problemática: {problematica or "—"}',
                        f'Acción: {accion or "—"}',
                        f'Equipo: {equipo or "—"}',
                        f'Plazo: {_plazo_texto(plazo_horas) or "—"}',
                        f'Ubicación (lat, lng): {coords}',
                    ]
                    _fund_any = False
                    for _p in (prob_in if isinstance(prob_in, list) else [problematica]):
                        _f = FUNDAMENTOS.get(str(_p).split(':')[0].strip())
                        if _f:
                            meta.append(f'   Fundamento {str(_p).strip()}: {_f["art"]} [{_f["tag"]}]')
                            _fund_any = True
                    if _fund_any:
                        meta.append('   [DS] R. Desarrollo Sustentable  ·  [Limp] R. Limpieza')
                    try:
                        foto_pdf = _fotos_a_pdf(fotos_dec, meta)
                    except Exception:
                        foto_pdf = None  # si algo falla con el PDF, no impedimos el registro
                cur.execute('''
                    INSERT INTO domicilios
                    (folio,folio_acta,fecha,direccion,uso,nombre_comercio,estado,problematica,
                     accion,equipo,plazo_horas,lat,lng,obs,foto_pdf)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                ''', (folio, folio_acta, fecha, direccion, uso, nombre_comercio, estado,
                      problematica, accion, equipo, plazo_horas, lat, lng, obs,
                      psycopg2.Binary(foto_pdf) if foto_pdf else None))
                dom_id = cur.fetchone()['id']
    finally:
        conn.close()
    return jsonify({'ok': True, 'folio': folio, 'id': dom_id}), 201

@app.route('/api/domicilios/<int:rid>/pdf', methods=['GET'])
@requiere_admin
def get_domicilio_pdf(rid):
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT folio, foto_pdf FROM domicilios WHERE id=%s", (rid,))
                row = cur.fetchone()
    finally:
        conn.close()
    if not row or not row['foto_pdf']:
        return jsonify({'error': 'Este domicilio no tiene fotografías.'}), 404
    data = row['foto_pdf']
    if isinstance(data, memoryview):
        data = data.tobytes()
    inline = bool(request.args.get('inline'))
    return send_file(io.BytesIO(bytes(data)), as_attachment=(not inline),
                     download_name=f"{row['folio']}_fotos.pdf",
                     mimetype='application/pdf')

@app.route('/api/domicilios/<int:rid>', methods=['DELETE'])
@requiere_admin
def eliminar_domicilio(rid):
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM domicilios WHERE id=%s", (rid,))
                if cur.rowcount == 0:
                    return jsonify({'error': 'Domicilio no encontrado'}), 404
                # Recalcular folio_dom al número más alto existente + 1 (o 1 si no quedan)
                cur.execute("""
                    UPDATE config SET valor = (
                        SELECT COALESCE(MAX(SUBSTRING(folio FROM 5)::INTEGER), 0) + 1
                        FROM domicilios
                    ) WHERE clave = 'folio_dom'
                """)
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/dashboard/domicilios', methods=['GET'])
@requiere_tablero
def dashboard_domicilios():
    _today = str(date.today())
    desde = request.args.get('desde', _today)
    hasta = request.args.get('hasta', _today)
    for _s in (desde, hasta):
        try:
            datetime.strptime(_s, '%Y-%m-%d')
        except (ValueError, TypeError):
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if desde > hasta:
        return jsonify({'error': 'desde debe ser anterior o igual a hasta'}), 400
    equipo_filtro = request.args.get('equipo', '').strip()

    base_where = " WHERE fecha BETWEEN %s AND %s"
    params = [desde, hasta]
    if equipo_filtro:
        base_where += " AND equipo = %s"
        params.append(equipo_filtro)

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                def qa(sql, p=None):
                    cur.execute(sql, p or [])
                    return cur.fetchall()
                def q1(sql, p=None):
                    cur.execute(sql, p or [])
                    return cur.fetchone()

                total = q1("SELECT COUNT(*) c FROM domicilios" + base_where, params)['c']

                por_uso = {r['k']: r['c'] for r in qa(
                    "SELECT COALESCE(NULLIF(TRIM(uso),''),'Ambos') k, COUNT(*) c FROM domicilios"
                    + base_where + " GROUP BY 1", params)}
                por_estado = {r['k']: r['c'] for r in qa(
                    "SELECT COALESCE(NULLIF(TRIM(estado),''),'—') k, COUNT(*) c FROM domicilios"
                    + base_where + " GROUP BY 1", params)}
                por_accion = {r['k']: r['c'] for r in qa(
                    "SELECT COALESCE(NULLIF(TRIM(accion),''),'—') k, COUNT(*) c FROM domicilios"
                    + base_where + " GROUP BY 1", params)}
                por_equipo = [{'k': r['k'], 'c': r['c']} for r in qa(
                    "SELECT COALESCE(NULLIF(TRIM(equipo),''),'—') k, COUNT(*) c FROM domicilios"
                    + base_where + " GROUP BY 1 ORDER BY c DESC", params)]
                hist = [dict(r) for r in qa(
                    "SELECT fecha, COUNT(*) c FROM domicilios" + base_where
                    + " GROUP BY fecha ORDER BY fecha", params)]
                # Problemática puede tener varias opciones por domicilio (separadas por coma):
                # se cuentan por separado. Se trae multa/accion para el desglose de multas.
                prob_rows = qa("SELECT problematica, multa, accion FROM domicilios" + base_where, params)
                venc = q1(
                    "SELECT "
                    "COUNT(*) FILTER (WHERE NOT COALESCE(cumplido,false) AND lim IS NOT NULL AND lim < now()) v, "
                    "COUNT(*) FILTER (WHERE NOT COALESCE(cumplido,false) AND lim IS NOT NULL AND lim >= now()) p "
                    "FROM (SELECT cumplido, " + SQL_LIMITE + " lim FROM domicilios" + base_where + ") t", params)
                # Resultados de plazos (solo datos numéricos para el tablero del jefe)
                res_plazos = q1(
                    "SELECT "
                    "COUNT(*) FILTER (WHERE COALESCE(cumplido,false) AND NOT COALESCE(incumplimiento,false)) cumplidos, "
                    "COUNT(*) FILTER (WHERE COALESCE(cumplido,false) AND COALESCE(incumplimiento,false)) incumplimientos, "
                    # Con multa: revisados con multa (aplica a amonestaciones).
                    # Sin multa: TODO lo revisado (cumplido) que no llevó multa, para que
                    # Con multa + Sin multa = total revisados (incluye notificados).
                    "COUNT(*) FILTER (WHERE multa IS TRUE AND TRIM(COALESCE(accion,''))='Amonestado') con_multa, "
                    "COUNT(*) FILTER (WHERE COALESCE(cumplido,false) AND NOT (multa IS TRUE AND TRIM(COALESCE(accion,''))='Amonestado')) sin_multa, "
                    "COUNT(*) FILTER (WHERE COALESCE(canalizado_ingresos,false)) canalizados "
                    "FROM domicilios" + base_where, params)
                pcfg = {r['clave']: r['valor'] for r in qa(
                    "SELECT clave, valor FROM config WHERE clave IN ('poligonos_asignados','poligonos_cubiertos','manzanas_por_poligono','manzanas_cubiertas')")}
    finally:
        conn.close()

    prob_counts = {}
    multa_prob_counts = {}
    for r in prob_rows:
        # La multa aplica solo a amonestaciones (mismo criterio que el desglose Con/Sin multa)
        es_multa = (r.get('multa') is True) and (str(r.get('accion') or '').strip() == 'Amonestado')
        partes = [p.strip() for p in (r['problematica'] or '').split(',') if p.strip()]
        for p in partes:
            prob_counts[p] = prob_counts.get(p, 0) + 1
        # Multas: una por caso, por su problemática principal (la primera del folio),
        # para que el total cuadre con el número de multas.
        if es_multa and partes:
            principal = partes[0]
            multa_prob_counts[principal] = multa_prob_counts.get(principal, 0) + 1
    por_problematica = sorted(
        [{'k': k, 'c': v} for k, v in prob_counts.items()],
        key=lambda x: -x['c'])
    multas_por_problematica = sorted(
        [{'k': k, 'c': v} for k, v in multa_prob_counts.items()],
        key=lambda x: -x['c'])

    return jsonify({
        'total': total,
        'por_uso': por_uso,
        'por_estado': por_estado,
        'por_problematica': por_problematica,
        'multas_por_problematica': multas_por_problematica,
        'por_accion': por_accion,
        'por_equipo': por_equipo,
        'historial': hist,
        'equipos': list(EQUIPOS),
        'vencidos': venc.get('v', 0) if venc else 0,
        'por_vencer': venc.get('p', 0) if venc else 0,
        'plazos_cumplidos': res_plazos.get('cumplidos', 0) if res_plazos else 0,
        'plazos_incumplimientos': res_plazos.get('incumplimientos', 0) if res_plazos else 0,
        'plazos_con_multa': res_plazos.get('con_multa', 0) if res_plazos else 0,
        'plazos_sin_multa': res_plazos.get('sin_multa', 0) if res_plazos else 0,
        'plazos_canalizados': res_plazos.get('canalizados', 0) if res_plazos else 0,
        'poligonos': _poligonos_payload(pcfg),
    })

@app.route('/api/tablero/domicilios', methods=['GET'])
@requiere_tablero
def tablero_domicilios():
    """Lista de los domicilios más recientes para el tablero de solo lectura."""
    _today = str(date.today())
    desde = request.args.get('desde', _today)
    hasta = request.args.get('hasta', _today)
    for _s in (desde, hasta):
        try:
            datetime.strptime(_s, '%Y-%m-%d')
        except (ValueError, TypeError):
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if desde > hasta:
        return jsonify({'error': 'desde debe ser anterior o igual a hasta'}), 400
    equipo = request.args.get('equipo', '').strip()
    try:
        limit = min(max(int(request.args.get('limit', 50)), 1), 200)
    except (ValueError, TypeError):
        limit = 50
    where = " WHERE fecha BETWEEN %s AND %s"
    params = [desde, hasta]
    if equipo:
        where += " AND equipo = %s"
        params.append(equipo)
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, folio, folio_acta, fecha, direccion, uso, nombre_comercio, "
                    "estado, problematica, accion, equipo, plazo_horas, "
                    "(foto_pdf IS NOT NULL) AS has_pdf, creado_en "
                    "FROM domicilios" + where + " ORDER BY id DESC LIMIT %s",
                    params + [limit])
                rows = cur.fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/tablero')
@requiere_tablero
def tablero():
    """Tablero de solo lectura (dirección). Sin formulario de captura."""
    resp = make_response(render_template('tablero.html'))
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    return resp

@app.route('/api/plazos', methods=['GET'])
@requiere_admin
def api_plazos():
    """Domicilios con plazo, para el módulo de seguimiento. estado_plazo:
    'vencido' | 'por_vencer' | 'cumplido'. Filtros opcionales: equipo y estado."""
    equipo = request.args.get('equipo', '').strip()
    estado = request.args.get('estado', 'pendientes')  # pendientes|vencidos|por_vencer|cumplidos|todos
    where = "WHERE plazo_horas IS NOT NULL"
    params = []
    if equipo:
        where += " AND equipo = %s"; params.append(equipo)
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, folio, folio_acta, fecha, direccion, uso, nombre_comercio, "
                    "estado, problematica, accion, equipo, plazo_horas, obs, "
                    "(foto_pdf IS NOT NULL) AS has_pdf, "
                    "COALESCE(cumplido,false) AS cumplido, cumplido_en, cumplido_obs, cumplido_por, multa, "
                    "COALESCE(canalizado_ingresos,false) AS canalizado_ingresos, "
                    "COALESCE(incumplimiento,false) AS incumplimiento, "
                    + SQL_LIMITE + " AS limite, "
                    "(" + SQL_LIMITE + " IS NOT NULL AND " + SQL_LIMITE + " < now()) AS vencido "
                    "FROM domicilios " + where + " ORDER BY folio ASC",
                    params)
                rows = cur.fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        d = dict(r)
        if d['cumplido']:
            d['estado_plazo'] = 'incumplimiento' if d.get('incumplimiento') else 'cumplido'
        else:
            d['estado_plazo'] = 'vencido' if d['vencido'] else 'por_vencer'
        out.append(d)
    if estado in ('vencidos', 'por_vencer', 'cumplidos', 'incumplimientos'):
        objetivo = {'vencidos': 'vencido', 'por_vencer': 'por_vencer',
                    'cumplidos': 'cumplido', 'incumplimientos': 'incumplimiento'}[estado]
        out = [d for d in out if d['estado_plazo'] == objetivo]
    elif estado == 'pendientes':
        out = [d for d in out if d['estado_plazo'] not in ('cumplido', 'incumplimiento')]
    return jsonify(out)

@app.route('/api/domicilios/<int:rid>/cumplir', methods=['POST'])
@requiere_admin
def cumplir_domicilio(rid):
    """Marca un domicilio como cumplido (plazo atendido) con observaciones. Solo admin."""
    d = request.get_json(silent=True) or {}
    obs = (d.get('obs') or '').strip()
    multa = d.get('multa')
    multa = bool(multa) if multa is not None else None
    canalizar = bool(d.get('canalizar'))
    # resultado: 'incumplimiento' o 'cumplido' (por defecto). Ambos cierran el plazo.
    incumplimiento = (d.get('resultado') == 'incumplimiento') or bool(d.get('incumplimiento'))
    por = (request.authorization.username if request.authorization else '') or ''
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE domicilios SET cumplido=TRUE, cumplido_en=now(), "
                    "cumplido_obs=%s, cumplido_por=%s, multa=%s, incumplimiento=%s, "
                    "canalizado_ingresos=%s, "
                    "canalizado_en = CASE WHEN %s THEN now() ELSE NULL END, "
                    "canalizado_por = CASE WHEN %s THEN %s ELSE NULL END "
                    "WHERE id=%s",
                    (obs, por, multa, incumplimiento, canalizar, canalizar, canalizar, por, rid))
                if cur.rowcount == 0:
                    return jsonify({'error': 'Domicilio no encontrado'}), 404
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/canalizados', methods=['GET'])
@requiere_admin
def api_canalizados():
    """Predios canalizados a Ingresos (por deshabitado). Solo admin. Orden por folio."""
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, folio, folio_acta, fecha, direccion, uso, nombre_comercio, "
                    "estado, problematica, equipo, lat, lng, "
                    "(foto_pdf IS NOT NULL) AS has_pdf, "
                    "canalizado_en, canalizado_por, cumplido_obs "
                    "FROM domicilios WHERE canalizado_ingresos = TRUE ORDER BY folio ASC")
                rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()
    return jsonify(rows)

@app.route('/api/export/excel/canalizados', methods=['GET'])
@requiere_admin
def export_excel_canalizados():
    """Excel de predios canalizados a Ingresos. Solo admin."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT folio, folio_acta, fecha, direccion, uso, nombre_comercio, "
                    "estado, problematica, equipo, lat, lng, "
                    "canalizado_en, canalizado_por, cumplido_obs "
                    "FROM domicilios WHERE canalizado_ingresos = TRUE ORDER BY folio ASC")
                rows = cur.fetchall()
    finally:
        conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Canalizados a Ingresos'
    headers = ['Folio captura', 'Folio acta física', 'Fecha inspección', 'Dirección',
               'Uso del predio', 'Nombre del comercio', 'Estado del predio',
               'Problemática', 'Equipo', 'Ubicación (lat, lng)',
               'Canalizado el', 'Canalizado por', 'Observación / resolución']
    header_fill = PatternFill(fill_type='solid', fgColor='1a6fc4')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for r in rows:
        coords = f"{r['lat']}, {r['lng']}" if (r.get('lat') is not None and r.get('lng') is not None) else ''
        ws.append([
            r['folio'], r.get('folio_acta', ''), r['fecha'], r['direccion'],
            r['uso'], r.get('nombre_comercio', ''), r['estado'],
            r['problematica'], r.get('equipo', ''), coords,
            r['canalizado_en'].strftime('%Y-%m-%d %H:%M') if r.get('canalizado_en') else '',
            r.get('canalizado_por', '') or '', r.get('cumplido_obs', '') or ''
        ])
    col_widths = [12, 14, 14, 34, 16, 24, 18, 22, 12, 22, 18, 16, 34]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='canalizados_ingresos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/reporte/domicilios.pdf', methods=['GET'])
@requiere_tablero
def reporte_domicilios_pdf():
    """Reporte informativo en PDF (resumen + avance de polígonos)."""
    _today = str(date.today())
    desde = request.args.get('desde', _today); hasta = request.args.get('hasta', _today)
    for _s in (desde, hasta):
        try:
            datetime.strptime(_s, '%Y-%m-%d')
        except (ValueError, TypeError):
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if desde > hasta:
        return jsonify({'error': 'desde debe ser anterior o igual a hasta'}), 400
    equipo = request.args.get('equipo', '').strip()
    where = " WHERE fecha BETWEEN %s AND %s"; params = [desde, hasta]
    if equipo:
        where += " AND equipo = %s"; params.append(equipo)
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT equipo, uso, problematica, accion, "
                    "COALESCE(cumplido,false) cumplido, COALESCE(incumplimiento,false) incumplimiento, "
                    "multa, COALESCE(canalizado_ingresos,false) canalizado_ingresos "
                    "FROM domicilios" + where, params)
                rows = [dict(r) for r in cur.fetchall()]
                # Amonestaciones por verificar = plazos vencidos sin cerrar (mismo criterio que el tablero)
                cur.execute(
                    "SELECT COUNT(*) v FROM (SELECT cumplido, " + SQL_LIMITE
                    + " lim FROM domicilios" + where + ") t "
                    "WHERE NOT COALESCE(cumplido,false) AND lim IS NOT NULL AND lim < now()", params)
                _venc = cur.fetchone()['v']
                cur.execute("SELECT clave, valor FROM config WHERE clave IN ('poligonos_asignados','poligonos_cubiertos','manzanas_por_poligono','manzanas_cubiertas')")
                pcfg = {r['clave']: r['valor'] for r in cur.fetchall()}
    finally:
        conn.close()
    # Conteos de seguimiento (numéricos) para el PDF
    seguimiento = {'vencidos': _venc, 'cumplidos': 0, 'incumplimientos': 0,
                   'con_multa': 0, 'sin_multa': 0, 'canalizados': 0}
    for r in rows:
        amon = str(r.get('accion') or '').strip() == 'Amonestado'
        if r['cumplido'] and not r['incumplimiento']:
            seguimiento['cumplidos'] += 1
        if r['cumplido'] and r['incumplimiento']:
            seguimiento['incumplimientos'] += 1
        if r.get('multa') is True and amon:
            seguimiento['con_multa'] += 1
        # Sin multa: todo lo revisado que no llevó multa (con_multa + sin_multa = revisados)
        if r['cumplido'] and not (r.get('multa') is True and amon):
            seguimiento['sin_multa'] += 1
        if r['canalizado_ingresos']:
            seguimiento['canalizados'] += 1
    _pol = _poligonos_payload(pcfg)
    _MES = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
    def _fmt(s):
        x = datetime.strptime(s, '%Y-%m-%d'); return f'{x.day} de {_MES[x.month-1]} de {x.year}'
    if desde <= '2020-01-01':
        fecha_txt = 'Acumulado total · al ' + _fmt(hasta)
    elif desde == hasta:
        fecha_txt = _fmt(desde)
    else:
        fecha_txt = _fmt(desde) + ' al ' + _fmt(hasta)
    from reporte import construir_reporte
    pdf = construir_reporte(rows, _pol['asignados'], POLY_COLORS, fecha_txt, _pol['manzanas'], _pol['cubiertas'], seguimiento=seguimiento)
    return send_file(io.BytesIO(pdf), mimetype='application/pdf', as_attachment=True,
                     download_name='Reporte_Operativo_' + desde + '.pdf')

@app.route('/api/operativo/poligonos', methods=['GET'])
@requiere_admin
def operativo_poligonos():
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT clave, valor FROM config WHERE clave LIKE 'poligonos_%%' OR clave LIKE 'manzanas_%%'")
                pcfg = {r['clave']: r['valor'] for r in cur.fetchall()}
    finally:
        conn.close()
    return jsonify(_poligonos_payload(pcfg))

@app.route('/api/operativo/manzanas', methods=['POST'])
@requiere_admin
def set_manzanas_cubiertas():
    d = request.get_json(silent=True) or {}
    pol = str(d.get('poligono', '')).strip()
    try:
        val = int(d.get('cubiertas'))
    except (TypeError, ValueError):
        return jsonify({'error': 'cubiertas debe ser un número'}), 400
    if not pol:
        return jsonify({'error': 'Falta el polígono'}), 400
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT clave, valor FROM config WHERE clave IN ('poligonos_cubiertos','manzanas_por_poligono','manzanas_cubiertas')")
                pcfg = {r['clave']: r['valor'] for r in cur.fetchall()}
                mz = json.loads(pcfg['manzanas_por_poligono']) if pcfg.get('manzanas_por_poligono') else MANZANAS_DEFAULT
                total = int(mz.get(pol, 0))
                if total <= 0:
                    return jsonify({'error': 'Polígono desconocido'}), 400
                val = max(0, min(val, total))
                if pcfg.get('manzanas_cubiertas'):
                    mc = json.loads(pcfg['manzanas_cubiertas'])
                else:
                    cub_list = json.loads(pcfg['poligonos_cubiertos']) if pcfg.get('poligonos_cubiertos') else POLY_CUBIERTOS
                    mc = {str(p): mz.get(str(p), 0) for p in cub_list}
                mc[pol] = val
                cur.execute("INSERT INTO config VALUES ('manzanas_cubiertas', %s) ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor", [json.dumps(mc)])
    finally:
        conn.close()
    return jsonify({'ok': True, 'poligono': pol, 'cubiertas': val, 'total': total})

@app.route('/api/operativo/mapa', methods=['GET'])
@requiere_tablero
def operativo_mapa():
    """Geometría de los polígonos (proyectada) para el mapa del tablero."""
    try:
        with open(os.path.join(os.path.dirname(__file__), 'mapa_poligonos.json'), encoding='utf-8') as f:
            geo = json.load(f)
    except Exception:
        geo = {'w': 0, 'h': 0, 'polys': []}
    geo['colores'] = POLY_COLORS
    resp = jsonify(geo)
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp

@app.route('/api/export/excel/domicilios', methods=['GET'])
@requiere_admin
def export_excel_domicilios():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT id, folio, folio_acta, fecha, direccion, uso, nombre_comercio,
                           estado, problematica, accion, equipo, plazo_horas, lat, lng, obs,
                           (foto_pdf IS NOT NULL) AS has_pdf, creado_en,
                           COALESCE(cumplido,false) AS cumplido, cumplido_en,
                           cumplido_obs, cumplido_por, multa,
                           COALESCE(incumplimiento,false) AS incumplimiento
                    FROM domicilios ORDER BY id DESC
                """)
                rows = cur.fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Domicilios'
    headers = ['ID', 'Folio captura', 'Folio acta física', 'Fecha', 'Dirección',
               'Uso del predio', 'Nombre del comercio', 'Estado del predio',
               'Problemática', 'Acción', 'Equipo', 'Plazo', 'Fecha límite',
               'Ubicación (lat, lng)', 'Fotos (PDF)', 'Observaciones', 'Registrado',
               'Cumplido', 'Resultado', 'Multa', 'Resolución', 'Confirmado por', 'Confirmado el']
    header_fill = PatternFill(fill_type='solid', fgColor='1a6fc4')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        lim = _fecha_limite(row['fecha'], row.get('plazo_horas'))
        coords = f"{row['lat']}, {row['lng']}" if (row.get('lat') is not None and row.get('lng') is not None) else ''
        ws.append([
            row['id'], row['folio'], row.get('folio_acta', ''), row['fecha'], row['direccion'],
            row['uso'], row.get('nombre_comercio', ''), row['estado'],
            row['problematica'], row.get('accion', ''), row.get('equipo', ''),
            _plazo_texto(row.get('plazo_horas')),
            lim.strftime('%Y-%m-%d %H:%M') if lim else '', coords,
            'Sí' if row.get('has_pdf') else 'No', row.get('obs', ''), row['creado_en'],
            'Sí' if row.get('cumplido') else 'No',
            ('Incumplimiento' if row.get('incumplimiento') else 'Cumplió') if row.get('cumplido') else '',
            ('Sí' if row.get('multa') else 'No') if row.get('multa') is not None else '',
            row.get('cumplido_obs', '') or '', row.get('cumplido_por', '') or '',
            row['cumplido_en'].strftime('%Y-%m-%d %H:%M') if row.get('cumplido_en') else ''
        ])
    col_widths = [6, 12, 14, 12, 34, 16, 24, 18, 22, 14, 12, 14, 18, 22, 10, 30, 18,
                  9, 14, 7, 30, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Domicilios_{date.today()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _colonias_para_mapa(rows):
    """Agrupa filas (norm, colonia, lat, lng, origen, c) por colonia:
    calcula el total de entradas y el origen dominante (el de mayor
    conteo) de cada una. Devuelve la lista ordenada por conteo desc."""
    por_colonia = {}
    for r in rows:
        if r['lat'] is None or r['lng'] is None:
            continue  # defensivo: fila sin coordenadas no debe tumbar el endpoint
        d = por_colonia.setdefault(r['norm'], {
            'colonia': r['colonia'], 'lat': float(r['lat']), 'lng': float(r['lng']),
            'total': 0, '_origenes': {}})
        c = int(r['c'])
        d['total'] += c
        o = r['origen'] or '—'
        d['_origenes'][o] = d['_origenes'].get(o, 0) + c
    salida = []
    for d in por_colonia.values():
        origen_dom = max(d['_origenes'].items(), key=lambda kv: kv[1])[0] if d['_origenes'] else '—'
        salida.append({'colonia': d['colonia'], 'lat': d['lat'], 'lng': d['lng'],
                       'origen': origen_dom, 'count': d['total']})
    salida.sort(key=lambda x: -x['count'])
    return salida

@app.route('/api/mapa', methods=['GET'])
@requiere_admin
def mapa_colonias():
    """Colonias de origen de las ENTRADAS del periodo con coordenadas
    (desde el caché colonias_geo), con su origen dominante. Solo aparecen
    las colonias que geocode_colonias.py pudo ubicar (estado='ok')."""
    _today = str(date.today())
    desde = request.args.get('desde', _today)
    hasta = request.args.get('hasta', _today)
    for _s in (desde, hasta):
        try:
            datetime.strptime(_s, '%Y-%m-%d')
        except (ValueError, TypeError):
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if desde > hasta:
        return jsonify({'error': 'desde debe ser anterior o igual a hasta'}), 400
    pga_filtro = request.args.get('pga', '').strip()

    params = [desde, hasta]
    pga_clause = ''
    if pga_filtro:
        pga_clause = ' AND r.pga = %s'
        params.append(pga_filtro)

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT LOWER(TRIM(r.colonia)) AS norm, MAX(TRIM(r.colonia)) AS colonia, "
                    "r.origen AS origen, COUNT(*) AS c, cg.lat AS lat, cg.lng AS lng "
                    "FROM registros r "
                    "JOIN colonias_geo cg ON cg.colonia_norm = LOWER(TRIM(r.colonia)) "
                    "WHERE r.tipo='ENTRADA' AND cg.estado='ok' "
                    "AND r.fecha BETWEEN %s AND %s "
                    "AND r.colonia IS NOT NULL AND TRIM(r.colonia) != ''"
                    + pga_clause +
                    " GROUP BY LOWER(TRIM(r.colonia)), r.origen, cg.lat, cg.lng",
                    params)
                rows = cur.fetchall()
    finally:
        conn.close()
    return jsonify({'colonias': _colonias_para_mapa(rows)})

@app.route('/api/dashboard', methods=['GET'])
@requiere_admin
def dashboard():
    _today = str(date.today())
    desde = request.args.get('desde', _today)
    hasta  = request.args.get('hasta',  _today)
    origen_filtro = request.args.get('origen', '').strip()
    pga_filtro = request.args.get('pga', '').strip()
    # Validate date format; fall back to today on bad input
    for _s in (desde, hasta):
        try:
            datetime.strptime(_s, '%Y-%m-%d')
        except (ValueError, TypeError):
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if desde > hasta:
        return jsonify({'error': 'desde debe ser anterior o igual a hasta'}), 400
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                def q(sql, p=None):
                    cur.execute(sql, p or [])
                    return cur.fetchone()
                def qa(sql, p=None):
                    cur.execute(sql, p or [])
                    return cur.fetchall()

                ent_periodo = q(
                    "SELECT COUNT(*) c FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='ENTRADA'",
                    [desde, hasta])['c']
                sal_periodo = q(
                    "SELECT COUNT(*) c FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='SALIDA'",
                    [desde, hasta])['c']
                tot    = q("SELECT COUNT(*) c FROM registros")['c']
                m3_ep  = q(
                    "SELECT COALESCE(SUM(m3),0) v FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='ENTRADA'",
                    [desde, hasta])['v']
                m3_sp  = q(
                    "SELECT COALESCE(SUM(m3),0) v FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='SALIDA'",
                    [desde, hasta])['v']
                m3_et  = q("SELECT COALESCE(SUM(m3),0) v FROM registros WHERE tipo='ENTRADA'")['v']
                m3_st  = q("SELECT COALESCE(SUM(m3),0) v FROM registros WHERE tipo='SALIDA'")['v']

                pgas = ['ROVIROSA WADE', 'LEY SAULO', 'ESTERITO', 'RECOVERDE', 'COMPRESORA']
                pga_flow = []
                for p in pgas:
                    ei  = q(
                        "SELECT COUNT(*) c FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='ENTRADA' AND pga=%s",
                        [desde, hasta, p])['c']
                    so  = q(
                        "SELECT COUNT(*) c FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='SALIDA'  AND pga=%s",
                        [desde, hasta, p])['c']
                    m3i = q(
                        "SELECT COALESCE(SUM(m3),0) v FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='ENTRADA' AND pga=%s",
                        [desde, hasta, p])['v']
                    m3o = q(
                        "SELECT COALESCE(SUM(m3),0) v FROM registros WHERE fecha BETWEEN %s AND %s AND tipo='SALIDA'  AND pga=%s",
                        [desde, hasta, p])['v']
                    pga_flow.append({
                        'pga': p, 'ent': ei, 'sal': so,
                        'm3_ent': round(float(m3i), 2),
                        'm3_sal': round(float(m3o), 2)
                    })

                origenes = ['NEGOCIO', 'RECOLECTORES', 'CASA-HABITACIÓN', 'CEA', 'LA OLA', 'CONTRATISTAS']
                origen_counts = []
                for o in origenes:
                    c = q("SELECT COUNT(*) c FROM registros WHERE tipo='ENTRADA' AND origen=%s", [o])['c']
                    if c:
                        origen_counts.append({'origen': o, 'count': c})
                otros = qa("""
                    SELECT origen, COUNT(*) c FROM registros
                    WHERE tipo='ENTRADA'
                      AND origen NOT IN ('NEGOCIO','RECOLECTORES','CASA-HABITACIÓN','CEA','LA OLA','CONTRATISTAS')
                      AND origen IS NOT NULL AND origen != '' AND origen != '—'
                    GROUP BY origen
                """)
                for r in otros:
                    origen_counts.append({'origen': r['origen'], 'count': r['c']})
                origen_counts.sort(key=lambda x: -x['count'])

                hist = qa("""
                    SELECT fecha,
                           SUM(CASE WHEN tipo='ENTRADA' THEN 1 ELSE 0 END) entradas,
                           SUM(CASE WHEN tipo='SALIDA'  THEN 1 ELSE 0 END) salidas
                    FROM registros
                    WHERE fecha BETWEEN %s AND %s
                    GROUP BY fecha ORDER BY fecha
                """, [desde, hasta])

                # Personas frecuentes: sin filtro de origen usa umbral > 3;
                # con un origen específico baja a > 1 (cada persona reparte sus
                # entradas entre varios orígenes, así que el conteo por origen es menor).
                # El umbral es un entero de nuestro propio código (1 o 3), no entrada
                # del usuario; el origen sí va como parámetro para evitar inyección.
                _nf_params = [desde, hasta]
                _nf_clauses = ''
                _nf_umbral = 3
                if origen_filtro:
                    _nf_clauses += ' AND origen = %s'
                    _nf_params.append(origen_filtro)
                    _nf_umbral = 1
                if pga_filtro:
                    _nf_clauses += ' AND pga = %s'
                    _nf_params.append(pga_filtro)
                # mode() WITHIN GROUP = valor más frecuente por persona:
                # su origen dominante y el PGA donde más dispuso (ignora NULLs,
                # por eso NULLIF de origen vacío).
                nombres_frecuentes = qa(
                    "SELECT (ARRAY_AGG(nombre ORDER BY id DESC))[1] AS nombre, COUNT(*) AS c, "
                    "mode() WITHIN GROUP (ORDER BY NULLIF(TRIM(origen),'')) AS origen, "
                    "mode() WITHIN GROUP (ORDER BY pga) AS pga "
                    "FROM registros "
                    "WHERE tipo='ENTRADA' AND fecha BETWEEN %s AND %s"
                    + _nf_clauses +
                    " AND nombre IS NOT NULL AND TRIM(nombre) != '' "
                    "GROUP BY LOWER(TRIM(nombre)) "
                    "HAVING COUNT(*) > " + str(_nf_umbral) + " "
                    "ORDER BY c DESC",
                    _nf_params)
    finally:
        conn.close()

    return jsonify({
        'ent_periodo': ent_periodo, 'sal_periodo': sal_periodo,
        'balance': max(0, ent_periodo - sal_periodo), 'total': tot,
        'm3_ent_periodo': round(float(m3_ep), 2), 'm3_sal_periodo': round(float(m3_sp), 2),
        'm3_ent_tot': round(float(m3_et), 2), 'm3_sal_tot': round(float(m3_st), 2),
        'pga_flow': pga_flow,
        'origen_counts': origen_counts,
        'nombres_frecuentes': [{'nombre': r['nombre'], 'count': r['c'],
                                'origen': r['origen'], 'pga': r['pga']} for r in nombres_frecuentes],
        'historial': [dict(r) for r in hist]
    })

@app.route('/api/export/excel', methods=['GET'])
@requiere_admin
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT * FROM registros ORDER BY id DESC")
                rows = cur.fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros'

    headers = ['ID', 'Folio', 'Tipo', 'Fecha', 'Hora', 'PGA', 'Detalle/Carga',
               'Origen', 'Nombre', 'Teléfono', 'Calle', 'Colonia', 'Vehículo', 'Placa', 'm³', 'Observaciones', 'Registrado']
    header_fill = PatternFill(fill_type='solid', fgColor='1a6fc4')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append([
            row['id'], row['folio'], row['tipo'], row['fecha'], row['hora'],
            row['pga'], row['detalle'], row['origen'], row['nombre'],
            row.get('telefono', ''), row['calle'], row['colonia'],
            row['vehiculo'], row['placa'], row['m3'], row['obs'], row['creado_en']
        ])

    col_widths = [6, 10, 8, 12, 8, 18, 18, 16, 22, 14, 18, 18, 18, 12, 8, 30, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"CT_Viajes_{date.today()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    init_db()
    app.run()
